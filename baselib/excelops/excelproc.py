import openpyxl
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
import os


class ExcelWorkBook:
    def __init__(self, workbook=None, worksheet=None, exists=True,
                 overwrite=True):
        self.wb = self._set_wb_object(workbook=workbook, overwrite=overwrite,
                                      exists=exists)
        self.ws = self._set_ws_object(worksheet=worksheet)

    def _set_wb_object(self, workbook, overwrite=True, exists=True):
        # o_workbook = None
        if os.path.isfile(workbook):
            if overwrite:
                os.remove(workbook)
        try:
            o_workbook = openpyxl.load_workbook(filename=workbook)
        except FileNotFoundError:
            o_workbook = openpyxl.Workbook()
        return o_workbook

    def _set_ws_object(self, worksheet):
        try:
            if worksheet in self.wb.sheetnames:
                return self.wb.get_sheet_by_name(worksheet)
            else:
                sheet = self.wb.active
                sheet.title = worksheet
                return sheet
        except AttributeError:
            sheet = self.wb.active
            sheet.title = worksheet
            return sheet

    def _add_first_row(self, data):
        colnumber = 1
        datadict = {}
        for key, value in data.items():
            self.ws.cell(row=1, column=colnumber).value = key
            colname = get_column_letter(colnumber)
            datadict[colname] = value
            colnumber += 1
        return datadict

    def _get_column_number(self, key, header=None):
        try:
            return header.index(key) + 1
        except ValueError:
            return self.ws.max_column + 1
        except AttributeError:
            return self.ws.max_column + 1

    def _add_first_row_new(self, data, header=None):
        try:
            assert isinstance(data, list)
            counter = 1
            for heads in header:
                self.ws.cell(row=1, column=counter).value = heads
                counter += 1
            # self.ws.append(header)
            return None
        except AssertionError:
            # colnumber = 1
            datadict = {}
            for key, value in data.items():
                colnumber = self._get_column_number(key, header=header)
                self.ws.cell(row=1, column=colnumber).value = key
                colname = get_column_letter(colnumber)
                datadict[colname] = value
            return datadict

    def _map_add_col_head_data(self, data, checkfirst=True, header=None):
        datadict = {}
        if checkfirst and self.ws.max_row < 2:
            # datadict = self._add_first_row(data=data)
            datadict = self._add_first_row_new(data=data, header=header)
        else:
            try:
                assert isinstance(data, list)
                datadict = None
            except AssertionError:
                for key, value in data.items():
                    match = 0
                    for colnum in range(1, self.ws.max_column + 1):
                        if self.ws.cell(row=1, column=colnum).value == key:
                            match = 1
                            break
                    if match == 1:
                        colname = get_column_letter(colnum)
                    else:
                        colnum = self._get_column_number(key, header=header)
                        self.ws.cell(row=1, column=colnum).value = key
                        colname = get_column_letter(colnum)
                    datadict[colname] = value
        return datadict

    def save_wb_objects(self, workbook):
        self.wb.save(workbook)

    def get_excel_columns(self, data, header=None):
        datalist = []
        recordnum = 0
        for tempdata in data:
            headwrite = True if recordnum == 0 else False
            if isinstance(tempdata, dict):
                datadict = self._map_add_col_head_data(tempdata, headwrite,
                                                       header=header)
                datalist.append(datadict)
            else:
                if headwrite:
                    datadict = self._map_add_col_head_data(tempdata, headwrite,
                                                           header=header)
                datalist.append(tempdata)
            recordnum += 1
        return datalist

    def append_to_ws(self, data):
        for tempdata in data:
            if isinstance(tempdata, list):
                self.ws.append(tempdata)
            elif isinstance(tempdata, dict):
                self.ws.append(tempdata)
            else:
                self.ws.append(data)
                break
