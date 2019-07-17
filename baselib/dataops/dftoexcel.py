import pandas as pd
import pandas.io.formats.excel as fmt_xl
from baselib.utils import objects as objs


class DFtoExcel:
    def __init__(self, **kwargs):
        self.engine = None
        self.replace = None
        self.startrow = None
        self.isclose = None
        self.datetimeformat = None
        self.dateformat = None
        self._set_dfexcelprop(**kwargs)
        self.xlswriter = None
        # self.backup_head_style = fmt_xl.header_style

    def _get_excel_writer(self, outfile):
        self.xlswriter = pd.ExcelWriter(outfile, engine=self.engine,
                                        datetime_format=self.datetimeformat,
                                        date_format=self.dateformat)
        return self.xlswriter

    def save_excel(self):
        if self.xlswriter is not None:
            self.xlswriter.save()
            self.xlswriter.close()

    def _set_dfexcelprop(self, **kwargs):
        dtfmt = "dd-mmm-yyyy hh:mm:ss"
        dfmt = dtfmt[:dtfmt.find(" ")].strip()
        self.engine = kwargs.get("engine", "xlsxwriter")
        self.replace = kwargs.get("replace", "False")
        self.startrow = kwargs.get("startrow", None)
        self.isclose = kwargs.get("close", True)
        self.datetimeformat = kwargs.get("datetimeformat", dtfmt)
        self.dateformat = kwargs.get("dateformat", dfmt)

    def _replace_inplace(self, filename, sheets, index=False):
        from openpyxl import load_workbook
        xlswriter = self._get_excel_writer(outfile=filename)
        startrow = self.startrow if self.startrow else 0
        try:
            wbook = load_workbook(filename=filename)
            xlswriter.book = wbook
            xlswriter.book = wbook
            for shtname in sorted(sheets.keys()):
                if self.startrow is None and \
                        shtname in xlswriter.book.sheetnames:
                    startrow = wbook[shtname].max_row
                if self.replace and shtname in wbook.sheetnames:
                    idx = wbook.sheetnames.index(shtname)
                    wbook.remove(wbook.worksheets[idx])
                    wbook.create_sheet(shtname, idx)
            xlswriter.sheets = dict((ws.title, ws) for ws in wbook.worksheets)
        except FileNotFoundError:
            pass
        for shtname in sorted(sheets.keys()):
            sheets[shtname].to_excel(xlswriter, shtname, startrow=startrow,
                                     index=index)
        if self.isclose:
            self.save_excel()

    def _write_to_excel(self, filename, sheets, index=False):
        xlswriter = self._get_excel_writer(outfile=filename)
        for shtname in sorted(sheets.keys()):
            sheets[shtname].to_excel(xlswriter, shtname, index=index)
        if self.isclose:
            self.save_excel()

    def write_to_excel(self, filename, sheets, inplace=False, index=False):
        if inplace:
            self._replace_inplace(filename=filename, sheets=sheets, index=index)
        else:
            self._write_to_excel(filename=filename, sheets=sheets, index=index)

    @staticmethod
    def readexcel(filename, sheets=None, **kwargs):
        retdict = {}
        if sheets:
            sheets = objs.return_list(sheets)
            for sheet in sheets:
                retdict[sheet] = pd.read_excel(filename, sheet_name=sheet,
                                               **kwargs)
        else:
            retdict = pd.read_excel(filename, sheet_name=None, **kwargs)
        return retdict

    @staticmethod
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        from openpyxl import load_workbook

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()
