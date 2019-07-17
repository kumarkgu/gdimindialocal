import re
from openpyxl import load_workbook
from baselib.utils import excelprocess as ep
from datetime import datetime


def split_name(name, regnumb=None, salutations=None, regsplit=None,
               regexp=None):
    first_name = ""
    middle_name = ""
    last_name = ""
    error_val = ""
    t_name = name.strip(" .\t")
    re_match = regnumb.search(t_name)
    if re_match:
        error_val = "'{}'. Value is Not a valid Name".format(t_name)
        return [name, first_name, middle_name, last_name, error_val]

    name_parts = regsplit.split(t_name)
    if len(name_parts) == 1:
        first_name = name_parts[0].strip(" .")
    else:
        is_salute = False
        for salute in salutations:
            t_prefix = salute
            if name_parts[0].lower() == salute.lower():
                if t_name.startswith(salute + "."):
                    t_prefix = salute + "."
                is_salute = True
                break
        fname_index = 0
        min_lname_index = 1
        min_mname_index = 3
        start_char = len(name_parts[0]) + 1
        if is_salute:
            error_val = "Possible Salutation: '{}'".format(
                name_parts[0].strip(" .")
            )
            fname_index += 1
            min_lname_index += 1
            min_mname_index += 1
            start_char = len(name_parts[0]) + len(name_parts[1]) + 2

        first_name = name_parts[fname_index].strip(" .")
        if len(name_parts) > min_lname_index:
            last_name = name_parts[-1].strip(" .")
            if len(name_parts) >= min_mname_index:
                length = len(t_name) - len(name_parts[-1])
                middle_name = t_name[start_char:length].strip(" .")
    return [name, first_name, middle_name, last_name, error_val]


def read_excel_file(filename=None, salutations=None, outfile=None):
    regexp = re.compile(r'^([^\s\.]+)[\s\.]*([^\s\.]*.*)[\s\.]+([^\s\.]+)$')
    re_numb = re.compile(r'[0-9]+[+-/]*')
    re_split = re.compile(r'[\.|\s][\s|\.]*')
    excelwb = load_workbook(filename=filename, read_only=True)
    sheet = excelwb.get_sheet_names()[0]
    excelws = excelwb.get_sheet_by_name(sheet)
    column = "H"
    data = []
    counter = 1
    multiplier = 1
    for row in range(2, excelws.max_row + 1):
        if counter == 500:
            currtime = datetime.now().strftime("%Y-%b-%d %H:%M:%S")
            print(
                "Time: {0}. Processed {1} records".format(
                    currtime,
                    str(500 * multiplier)
                )
            )
            if multiplier == 1:
                write_out_file(outfile=outfile, data=data, exceloverwrite=True)
            else:
                write_out_file(outfile=outfile, data=data, exceloverwrite=False)
            data = []
            multiplier += 1
            counter = 0
        cell_name = "{}{}".format(column, row)
        name = excelws[cell_name].value
        # print("..Processing Name: {}".format(name))
        data.append(
            split_name(
                name=name, regnumb=re_numb, salutations=salutations,
                regsplit=re_split, regexp=regexp
            )
        )
        counter += 1
    write_out_file(outfile=outfile, data=data, exceloverwrite=False)
    # return data


def append_to_excel(worksheet, data, header=None):
    try:
        assert isinstance(data, list)
        output = worksheet.get_excel_columns(data=data, header=header)
    except AssertionError:
        output = worksheet.get_excel_columns(data=[data])
    worksheet.append_to_ws(output)


def write_out_file(outfile, data, **kwargs):
    worksheet = ep.ExcelWorkBook(workbook=outfile, worksheet="Output",
                                 overwrite=kwargs.get('exceloverwrite', False))
    header = ["Full Name", "First Name", "Middle Name", "Last Name", "Error"]
    append_to_excel(worksheet=worksheet, data=data, header=header)
    worksheet.save_wb_objects(outfile)


def process_file():
    userpath = "C:/Users/gunjan.kumar"
    projpath = "{}/Documents/JLL/Projects".format(userpath)
    # files = ["Emergency_Contact_Sri_Lanka.xlsx",
    #          "Emergency_Contact_India.xlsx"]
    files = ["Emergency_Contact_India.xlsx"]
    # files = ["Test_Data.xlsx"]
    salutations = [
        "Ch", "Bi", "Dr", "MD", "Mr", "MS", "SR", "SH", "Mrs", "Sri", "Smt",
        "Baba", "Baby", "Miss", "Shri", "Sree", "Syed", "Shree", "Doctor"
    ]
    for file in files:
        filename = "{0}/HRAnalytics/PSData/{1}".format(projpath, file)
        outf = file.rsplit('.', 1)[0] + "_output.xlsx"
        outfile = "{0}/HRAnalytics/PSData/{1}".format(projpath, outf)
        currtime = datetime.now().strftime("%Y-%b-%d %H:%M:%S")
        print(
            "Processing File: {0}. Time: {1}".format(
                filename,
                currtime
            )
        )
        read_excel_file(filename=filename, salutations=salutations,
                        outfile=outfile)
        # write_out_file(outfile=outfile, data=ldata, exceloverwrite=True)


process_file()


# is_salute = False
# t_name = name
# for salute in salutations:
#     if name.startswith(salute):
#         t_prefix = salute + "."
#         if name.startswith(t_prefix):
#             t_name = name[len(t_prefix):]
#             break
#         t_name = name[len(salute):]
#         t_prefix = salute
#         is_salute = True
#         break
# name_parts = regsplit.split(t_name)
# if len(name_parts) == 1:
#
#
# re_match = regexp.search(name)
# if re_match:
#     first_name = re_match.group(1)
#     middle_name = re_match.group(2)
#     last_name = re_match.group(3)
#     if middle_name != "" and last_name == "":
#         last_name = middle_name
#         middle_name = ""
# else:
#     first_name = name
