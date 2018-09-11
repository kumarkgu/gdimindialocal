import re
import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pandas.api.types import is_numeric_dtype
from pandas.api.types import is_string_dtype
from base.utils import Logger as lo

# auditfile = "{}/{}.xlsx".format(
#      r"C:\Users\rachael.wang\Documents\JLL\Projects\JapanPdfToText\Files\dq",
#      "csv_dq"
# )
# pdffile = "MFBM_20180702"
# outfile = "{}/{}.csv".format(
#     r"C:\Users\rachael.wang\Documents\JLL\Projects\JapanPdfToText\Files\process\output\201805",
#     pdffile
# )
# auditout= "{}/{}.xlsx".format(
#      r"C:\Users\rachael.wang\Documents\JLL\Projects\JapanPdfToText\Files\process\output\201805\dq",
#      "dq"
# )

class csv_dq():

    def __init__(self, **kwargs):
        self.pdffile = None
        self.auditfile = None
        self.outfile = None
        self.processname = kwargs.get('processname', 'japan')
        self.log = kwargs.get('log', self._set_logger())




    @staticmethod
    def import_audit_sheet(auditfile, pdffile):
        """
        :param auditfile: location of the audit file
        :param pdffile: file name of the PDF file
        :return: dataframe of the auditfile for the pdffile
        """
        audit = pd.read_excel(auditfile)
        filename = pdffile.split('_')[0]
        audit = audit[audit["FileName"] == filename]
        return audit

    @staticmethod
    def import_outfile(outfile):
        """
        :param outfile: location of the output csv file
        :return: dataframe of the output csv file
        """
        outfiledf = pd.read_csv(outfile)
        outfiledf.columns = outfiledf.columns.str.replace(" ","")
        return outfiledf

    @staticmethod
    def write_to_excel(outfile, sheetname, df):
        writer = None
        if os.path.isfile(outfile):
            writer = pd.ExcelWriter(outfile, engine='openpyxl')
            writer.book = load_workbook(outfile)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            if sheetname in writer.book.sheetnames:
                startrow = writer.book[sheetname].max_row
                header = None
            else:
                startrow = 0
                header = 1
            df.to_excel(writer, sheetname, startrow=startrow , header = header, index=False)
        else:
            writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
            df.to_excel(writer, sheet_name=sheetname, index=False)
        writer.save()
        return None

    @staticmethod
    def check_header(audit, pdffile, outfiledf):
        """
        :param audit: dataframe of the auditfile
        :param outfiledf: dataframe of the output csv file
        :return: one row log for the header
        """
        filename = pdffile
        sourcol = [x.replace(" ","") for x in list(outfiledf.columns)]
        tarcol = [x.replace(" ","") for x in list(audit["Column"])]

        miss = []
        add = []

        for col in tarcol:
            if col not in sourcol:
                miss.append(col)

        for col in sourcol:
            if col not in tarcol:
                add.append(col)

        if miss != [] and add != []:
            return [filename, ",".join(miss), ",".join(add)]
        else: return None

    @staticmethod
    def check_col_dtype(sourcol, type):
        if type == "str":
            check = is_string_dtype(sourcol)
        elif type == "float":
            check = is_numeric_dtype(sourcol)
        elif type == "int":
            if sourcol.dtype == np.int64:
                check = True
            else:
                check = False
        elif "年" in type:
            check = False
        else: check = None
        return check

    @staticmethod
    def check_row_dtype(value, dtype):
        if type(value).__name__ != dtype and value is not None:
            badvalue = value
            reason = "Data Type should be " + dtype + ", but it's now " + type(value).__name__ + " in the csv file."
            return badvalue, reason
        else: return None, None

    @staticmethod
    def check_row_format(value, dtype):
        if not re.compile(dtype).match(value):
            badvalue = value
            reason = "Data format should be " + dtype + ", but it's " + badvalue + " in the csv file."
            return badvalue, reason
        else: return None, None

    @staticmethod
    def app_to_df(dqContentlog, filename, column, line, badvalue, reason):
        if not badvalue:
            return dqContentlog
        else:
            if dqContentlog is None:
                dqContentlog = [[filename, column, line, badvalue, reason]]
            else:
                dqContentlog.append([filename, column, line, badvalue, reason])
            return dqContentlog

    def check_content(self, audit, pdffile, outfiledf):
        """
        :param audit: dataframe of the auditfile
        :param outfiledf: dataframe of the output csv file
        :return: dataframe log for bad values
        """
        dqContentlog = None
        _filename = pdffile
        _columns = ["FileName", "Column", "Line", "BadValue", "Reason"]

        for index, row in audit.iterrows():
            i = 1
            _column = row["Column"]
            _dtype = row["DataType"]
            _sourcol = outfiledf[_column].dropna()

            _check = self.check_col_dtype(_sourcol, _dtype)

            if not _check:
                if _dtype in ("str","float","int"):
                    for value in _sourcol:
                        _badvalue, _reason = self.check_row_dtype(value, _dtype)
                        dqContentlog = self.app_to_df(dqContentlog, _filename, _column, i , _badvalue, _reason)
                        i = i + 1
                if "年" in _dtype:
                    for value in _sourcol:
                        _badvalue, _reason = self.check_row_format(value, _dtype)
                        dqContentlog = self.app_to_df(dqContentlog, _filename, _column, i , _badvalue, _reason)
                        i = i + 1
                dqContentlogdf = pd.DataFrame(dqContentlog,columns=_columns)
        return dqContentlogdf


    def dq_check(self, auditfile, pdffile, outfile, auditout):

        audit = self.import_audit_sheet(auditfile = auditfile, pdffile = pdffile)
        outfiledf = self.import_outfile(outfile = outfile)

        header_output = self.check_header(audit = audit, pdffile = pdffile, outfiledf = outfiledf)

        if header_output is None:
            self.log.info(
                "..[DQ check] - Header check on " + pdffile + " is completed - No error."
            )
            content_output = self.check_content( audit = audit, pdffile = pdffile, outfiledf = outfiledf)
            self.log.info(
                "..[DQ check] - Content check on " + pdffile + " is compeleted."
            )
            self.write_to_excel(auditout, "Header", content_output)
            return header_output, content_output
        else:
            self.log.info(
                "..[DQ check] - Header check on" + pdffile + " is completed - With Error"
            )
            self.write_to_excel(auditout, "Content", header_output)
            return header_output, None


#csv_dq = csv_dq()
#header_output, content_output = csv_dq.dq_check(auditfile=auditfile , pdffile= pdffile, outfile= outfile, auditout = auditout)

