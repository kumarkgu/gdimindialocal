import re
import os
import glob
import csv
from bs4 import BeautifulSoup
from base.pdf import xpdf as xpdf
from base.pdf import tabula as tabula
from base.utils import fileobjects as fo
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from pandas.api.types import is_numeric_dtype
from pandas.api.types import is_string_dtype
from base.utils import Logger as lo

class JapanBasePDF:
    def __init__(self, xpdfdir=None, tabuladir=None, tabulajarfile=None,
                 processdir=None, tempname="temp", outname="output",
                 dqname="dq", auditfile = None, **kwargs):
        self.tabdir = tabuladir
        self.tabjarfile = tabulajarfile
        self.xpdfdir = xpdfdir
        self.processdir = processdir
        self.tempdir = "{}/{}".format(processdir, tempname)
        self.outdir = "{}/{}".format(processdir, outname)
        self.dqoutdir = "{}/{}".format(self.outdir, dqname)
        self.auditfile = auditfile
        self.htmldir = "{}/html".format(processdir)
        self.regfile = re.compile(r'page(\d+)\.html')
        self.pxreg = re.compile(r'\D*(\d+)px', re.IGNORECASE)
        self.xpdf = xpdf.XpdfPdfProcess(utilpath=xpdfdir)
        self.tabula = tabula.TabulaPDF(tabuladir=tabuladir,
                                       tabulajarfile=tabulajarfile)
        self.processname = kwargs.get('processname', 'japan')
        self.log = kwargs.get('log', self._set_logger())

    def _set_logger(self):
        return lo.Logger(self.processname).getlog()

    def _exec_xpdf_pdftohtml(self, pdffile, htmldir):
        self.xpdf.pdf_to_html(pdffile, htmldir=htmldir)

    def _exec_tabula(self, pdffile, outfile, boundary=None, pageno=None,
                     runtype="lattice"):
        pages = 'all' if not pageno else pageno
        self.tabula.execute_tabula(
            pageno=pages,
            boundary=boundary,
            outfile=outfile,
            pdffile=pdffile,
            runtype=runtype
        )

    def _get_html_files(self, htmldir):
        templist = []
        filelist = []
        for file in glob.glob("{}/page*.html".format(htmldir)):
            basename = os.path.basename(file)
            regmatch = self.regfile.search(basename)
            if regmatch:
                templist.append(int(regmatch.group(1)))
        templist.sort()
        for fnumber in templist:
            filelist.append(
                "{0}/page{1}.html".format(htmldir, str(fnumber))
            )
        return filelist

    @staticmethod
    def _get_top_position(divs, regexp):
        posn = 0
        for keyval in divs["style"].split(";"):
            key = keyval.strip().split(":")
            if key[0] == "top":
                regmatch = regexp.search(key[1])
                if regmatch:
                    posn = regmatch.group(1)
                    break
        return posn

    def _get_topbottom_coord(self, filename, parser=None, begin=None, end=None):
        bsoup = BeautifulSoup(open(filename, encoding='utf-8'), parser)
        beginposn = 0
        endposn = 0
        regmatch = self.regfile.search(filename)
        filenumber = regmatch.group(1)
        alldivs = bsoup.findAll('div', {'class': 'txt'})
        flag = 0
        for divs in alldivs:
            currtext = divs.getText()
            if begin and begin != "":
                flag = 1
            if end and end != "":
                flag += 10
            currposn = int(self._get_top_position(divs, self.pxreg))
            if flag > 0:
                if flag == 1 or flag == 11:
                    if currtext == begin:
                        beginposn = currposn
                    if flag == 1:
                        endposn = currposn if currposn > endposn else endposn
                if flag == 10 or flag == 11:
                    if currtext == end:
                        endposn = currposn
                    if flag == 10:
                        beginposn = currposn if currposn < beginposn else beginposn
                if flag == 11:
                    if beginposn > 0 and endposn > 0:
                        break
            else:
                beginposn = currposn if currposn < beginposn else beginposn
                endposn = currposn if currposn > endposn else endposn
        beginposn = beginposn - 5
        endposn = endposn + 35 if flag == 1 else endposn + 1
        return beginposn, endposn, filenumber

    @staticmethod
    def refill_csv(infile, outfile, ignorefirst=True):
        linecounter = 1
        outcsv = open(outfile, 'wt', encoding='utf-8', newline="")
        writer = csv.writer(outcsv, delimiter=",")
        prevline = []
        with open(infile, 'rt', encoding='utf-8') as incsv:
            reader = csv.reader(incsv, dialect='excel')
            for line in reader:
                outline = [field for field in line]
                if linecounter == 1:
                    linecounter += 1
                    if not ignorefirst:
                        writer.writerow(outline)
                    continue
                try:
                    for posn in range(0, len(outline)):
                        if outline[posn] == "":
                            outline[posn] = prevline[posn]
                        else:
                            if posn == 0:
                                outline[posn] = outline[posn].split(";")[0]
                except IndexError:
                    pass
                prevline = outline
                writer.writerow(outline)
                linecounter += 1
            outcsv.close()

    def build_directory(self, tempdir, outputdir):
        fo.create_dir(tempdir)
        fo.create_dir(outputdir)

    def preprocess_pdf(self, pdffile, html_dir=None):
        if not html_dir:
            htmldir = "{}/html".format(os.path.dirname(pdffile))
        else:
            htmldir = html_dir
        fo.remove_dir(htmldir)
        self._exec_xpdf_pdftohtml(pdffile=pdffile, htmldir=htmldir)
        filelist = self._get_html_files(htmldir=htmldir)
        return filelist

    def process_pdf_tab(self, pdffile, htmlfile, outfile, begin, end, **kwargs):
        topposn, bottomposn, pageno = self._get_topbottom_coord(
            htmlfile,
            "html.parser",
            begin=begin,
            end=end
        )
        xycoord = "{},{},{},{}".format(
            str(topposn),
            str(kwargs.get('left', 2)),
            str(bottomposn),
            str(kwargs.get('right', 1450))
            #str(kwargs.get('right', 850))
        )
        self._exec_tabula(pdffile=pdffile, outfile=outfile, boundary=xycoord,
                          pageno=pageno,
                          runtype=kwargs.get('runtype', 'lattice'))

    """
    Code below this line are for DQ check on csv output
    """
    @staticmethod
    def import_audit_sheet(auditfile, pdffile):
        """
        :param auditfile: location of the audit file
        :param pdffile: file name of the PDF file with Path
        :return: dataframe of the auditfile for the pdffile
        """
        basefile = fo.get_base_filename(pdffile)
        audit = pd.read_excel(auditfile)
        filename = basefile.split('_')[0]
        audit = audit[audit["FileName"] == filename]
        return audit, basefile

    @staticmethod
    def csv_to_df(csvpath):
        """
        :param outfile: location of the output csv file
        :return: dataframe of the output csv file
        """
        header = pd.read_csv(csvpath
                             ,error_bad_lines=False
                             ,sep=","
                             ,index_col=False
                             ,encoding="utf-8"
                             ,nrows=0)
        df = pd.read_csv(csvpath
                         ,error_bad_lines=False
                         ,sep=","
                         ,index_col=False
                         ,encoding="utf-8"
                         ,skiprows=1
                         ,names=header.columns)
        df.columns = df.columns.str.replace(" ","")
        return df

    @staticmethod
    def write_to_excel(outfile, sheetname, df):
        writer = None
        if os.path.isfile(outfile):
            writer = pd.ExcelWriter(outfile, engine='openpyxl')
            writer.book = load_workbook(outfile)
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
            if sheetname in writer.book.sheetnames:
                startrow = writer.book[sheetname].max_row
                header = None
            else:
                startrow = 0
                header = 1
            df.to_excel(writer, sheetname, startrow=startrow , header = header, index=False)
        else:
            writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
            df.to_excel(writer, sheet_name=sheetname, index=False)
        writer.save()
        return None

    @staticmethod
    def check_header(audit, basefile, outfiledf):
        """
        :param audit: dataframe of the auditfile
        :param outfiledf: dataframe of the output csv file
        :return: one row log for the header
        """
        _filename = basefile
        _sourcol = [x.replace(" ", "") for x in list(outfiledf.columns)]
        _tarcol = [x.replace(" ", "") for x in list(audit["Column"])]
        _columns = ["FileName", "Missing Column", "Additional Column", "Action"]

        miss = []
        add = []

        for col in _tarcol:
            if col not in _sourcol:
                miss.append(col)

        for col in _sourcol:
            if col not in _tarcol:
                add.append(col)

        if miss != [] and add != []:
            return pd.DataFrame({"FileName":[_filename],
                                "Missing Column":[",".join(miss)],
                                "Additional Column": [",".join(add)],
                                "Action": ["Please ensure no missing column."]},
                                columns = _columns)
        else:
            return None

    @staticmethod
    def check_col_dtype(sourcol, type):
        if type == "str":
            check = is_string_dtype(sourcol)
        elif type == "float":
            check = is_numeric_dtype(sourcol)
        elif type == "int":
            if sourcol.dtype == np.int64:
                check = True
            else:
                check = False
        elif "年" in type:
            check = False
        else:
            check = None
        return check

    @staticmethod
    def check_row_dtype(value, dtype):

        _value = value
        if dtype == "float":
            try:
                _value = float(value.replace(",",""))
            except:
                _value = value

        if dtype == "int":
            try:
                _value = int(value)
            except:
                _value = value

        if type(_value).__name__ != dtype and _value is not None:
            badvalue = value
            reason = "Data Type should be " + dtype + ", but it's now " + type(_value).__name__ + " in the csv file."
            return badvalue, reason
        else:
            return None, None

    @staticmethod
    def check_row_format(value, dtype):
        if not re.compile(dtype).match(value) and value is not None:
            badvalue = value
            reason = "Data format should be " + dtype + ", but it's " + badvalue + " in the csv file."
            return badvalue, reason
        else:
            return None, None

    @staticmethod
    def app_to_df(dqContentlog, filename, column, line, badvalue, reason):
        if not badvalue:
            return dqContentlog
        else:
            if dqContentlog is None:
                dqContentlog = [[filename, column, line, badvalue, reason]]
            else:
                dqContentlog.append([filename, column, line, badvalue, reason])
            return dqContentlog

    def check_content(self, audit, basefile, outfiledf):
        """
        :param audit: dataframe of the auditfile
        :param outfiledf: dataframe of the output csv file
        :return: dataframe log for bad values
        """
        dqContentlog = None
        dqContentlogdf = pd.DataFrame()
        _filename = basefile
        _columns = ["FileName", "Column", "Line", "BadValue", "Reason"]

        for index, row in audit.iterrows():
            i = 1
            _column = row["Column"]
            _dtype = row["DataType"]
            _sourcol = outfiledf[_column].dropna()

            _check = self.check_col_dtype(_sourcol, _dtype)

            if not _check:
                if _dtype in ("str", "float", "int"):
                    for value in _sourcol:
                        _badvalue, _reason = self.check_row_dtype(value, _dtype)
                        dqContentlog = self.app_to_df(dqContentlog, _filename, _column, i, _badvalue, _reason)
                        i = i + 1
                if "年" in _dtype:
                    for value in _sourcol:
                        _badvalue, _reason = self.check_row_format(value, _dtype)
                        dqContentlog = self.app_to_df(dqContentlog, _filename, _column, i, _badvalue, _reason)
                        i = i + 1
                dqContentlogdf = pd.DataFrame(dqContentlog, columns=_columns)

        return dqContentlogdf

    @staticmethod
    def check_null(audit, basefile, outfiledf):
        _filename = basefile
        _columns = ["FileName", "ColumnName", "NullPercent", "Action"]
        null_output = pd.DataFrame(columns=_columns)
        for index, row in audit.iterrows():
            _column = row["Column"]
            _nullability = row["Nullability"]
            _sourcol = outfiledf[_column]

            if _nullability == "Not Null":
                if _sourcol.replace(" ","").replace("",np.nan).isnull().values.any():
                    nullpercent = _sourcol.replace(" ","").replace("",np.nan).isnull().sum()/ len(_sourcol)
                    null_output = null_output.append(pd.DataFrame({"FileName":[_filename],
                                                                   "ColumnName":[_column],
                                                                   "NullPercent": [nullpercent],
                                                                   "Action": ["This column should not be null, please check the csv output."]},
                                                                  columns = _columns))

        return null_output

    def dq_check(self, auditfile, pdffile, outfile, auditout):

        audit, basefile  = self.import_audit_sheet(auditfile=auditfile, pdffile=pdffile)
        outfiledf = self.csv_to_df(csvpath = outfile)

        header_output = self.check_header(audit=audit, basefile=basefile, outfiledf=outfiledf)

        if header_output is None:
            self.log.info(
                "..[DQ check] - Header check on " + basefile + " is completed - No error."
            )
            content_output = self.check_content(audit=audit, basefile=basefile, outfiledf=outfiledf)
            null_output = self.check_null(audit = audit, basefile = basefile, outfiledf = outfiledf)
            if content_output is None and null_output is None:
                self.log.info(
                    "..[DQ check] - Content check on " + basefile + " is compeleted - No error."
                )
            else:
                self.log.info(
                    "..[DQ check] - Content check on " + basefile + " is compeleted - With error."
                )
                if not content_output.empty:
                    self.write_to_excel(auditout, "Datatype", content_output)
                if not null_output.empty:
                    self.write_to_excel(auditout, "Nullability", null_output)
            return header_output, content_output
        else:
            self.log.info(
                "..[DQ check] - Header check on" + basefile + " is completed - With Error"
            )
            self.write_to_excel(auditout, "Header", header_output)
            return header_output, None


